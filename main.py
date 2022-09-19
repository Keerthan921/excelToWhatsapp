import openpyxl as xl
import pywhatkit as pk
from datetime import datetime

sourceFilename = "Notification List.xlsx"

loadedExcel = xl.load_workbook(sourceFilename)
NameListSheet = loadedExcel.worksheets[0]
TemplateSheet = loadedExcel.worksheets[1]

mr = NameListSheet.max_row
maxRowSheet2 = TemplateSheet.max_row

f = open("List.txt", "w")
f.write("|---------------------------------------------------|\n")
f.close()
for i in range(2, mr + 1):
    if NameListSheet.cell(row=i, column=7).value == 'Y':
        Name = NameListSheet.cell(row=i, column = 3).value
        PhoneNo = NameListSheet.cell(row=i, column=4).value
        EventType = NameListSheet.cell(row=i, column=1).value
        EventName = NameListSheet.cell(row=i, column=2).value
        date_str = str(NameListSheet.cell(row=i, column=5).value)
        date_str = date_str.strip(' 00:00:00')
        date_str = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%y')
        time_str = str(NameListSheet.cell(row=i, column=6).value)
        time_str = datetime.strptime(time_str, '%H:%M:%S').strftime('%I:%M %p')

        for j in range(2, maxRowSheet2 + 1):
            if TemplateSheet.cell(row=j, column=1).value == EventType:
                templateMessage = TemplateSheet.cell(row=j, column=2).value
            if TemplateSheet.cell(row=j, column=1).value == 'First Message':
                introMessage = TemplateSheet.cell(row=j, column=2).value

        dataToFillTemplate = {'[EventType]': EventType,
                              '[EventName]': EventName,
                              '[Date]': date_str,
                              '[Time]': time_str}

        for key, value in dataToFillTemplate.items():
            templateMessage = templateMessage.replace(key, value)

        
        nameHeader = (Name + "  "+ str(PhoneNo) +"\n\n\n")
                                                                    

        print(introMessage)
        print(templateMessage)
        pk.sendwhatmsg_instantly("+65" + str(PhoneNo), introMessage, 10, True, 15)
        pk.sendwhatmsg_instantly("+65" + str(PhoneNo), templateMessage, 10, True, 15)
        f = open("List.txt", "a")
        f.write(nameHeader)
       
        f.write(introMessage)
        f.write(templateMessage)
        f.write("\n|---------------------------------------------------|\n")

        now = datetime.now()
        TimeNow = now.strftime("%d/%m/%Y %H:%M")
        NameListSheet.cell(row=i, column=8).value = TimeNow

loadedExcel.save('Notification List.xlsx')


